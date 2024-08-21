from datetime import datetime
import pytz
import re
from telegram import Update, InputFile, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, filters, ContextTypes, CallbackQueryHandler
from openpyxl import Workbook, load_workbook
from texts import * #TEXTS, get_text, detect_language

# Ruta del archivo Excel
EXCEL_PATH = "datos_extraidos.xlsx"
CHATWARS_ID = "265204902"

###########################SECURITY###############################################################

def load_authorized_users(file_path):
    try:
        with open(file_path, 'r') as file:
            users = [int(line.strip()) for line in file.readlines()]
        return users
    except FileNotFoundError:
        print(f"Error: El archivo {file_path} no fue encontrado.")
        return []
    except ValueError:
        print("Error: Asegúrate de que todos los IDs en el archivo sean números enteros.")
        return []

def save_authorized_users(file_path, users):
    try:
        with open(file_path, 'w') as file:
            for user_id in users:
                file.write(f"{user_id}\n")
    except Exception as e:
        print(f"Error al guardar en {file_path}: {e}")

AUTHORIZED_USERS = load_authorized_users("users.txt")

def is_authorized(user_id):
    """Verifica si el usuario está autorizado."""
    return user_id in AUTHORIZED_USERS, "ANY"
    
async def validate(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if is_authorized(update.effective_user.id):
        if context.args:
            try:
                new_user_id = int(context.args[0])
                if new_user_id not in AUTHORIZED_USERS:
                    AUTHORIZED_USERS.append(new_user_id)
                    save_authorized_users("users.txt", AUTHORIZED_USERS)
                    await update.message.reply_text(f"El usuario con ID {new_user_id} ha sido validado exitosamente.")
                else:
                    await update.message.reply_text(f"El usuario con ID {new_user_id} ya está validado.")
            except ValueError:
                await update.message.reply_text("Por favor, proporciona un ID de usuario válido.")
        else:
            await update.message.reply_text("Por favor, proporciona el ID del usuario que deseas validar.")
    else:
        await update.message.reply_text("Lo siento, no tienes permiso para usar este bot.")

##################################################################################################

def cargar_o_crear_excel():
    """Carga el archivo Excel si existe, de lo contrario, lo crea."""
    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos Extraídos"
        ws.append(["Ubicación", "🇲🇴", "🇻🇦", "🇮🇲", "🇪🇺", "text", time, user_posted])  # Encabezados
        wb.save(EXCEL_PATH)
    return wb, ws

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    #####SEGURIDAD#####
    user_id = update.effective_user.id
    if is_authorized(user_id):
        await update.message.reply_text(get_text(update, 'welcome').format(name=update.effective_user.first_name)+"\n\n"+get_text(update, 'help_message'))
    else:
        await update.message.reply_text(get_text(update,'no_permission'))
    ###################

def extract_location(message: str):
    """Extrae la ubicación del mensaje y la convierte al formato esperado."""
    pattern = r"(?:0#0|([RGBY]{1,2})\s*(\d+)(?:#(\d+))?)"
    match = re.search(pattern, message)

    if match:
        color_prefix = match.group(1).lower()  # Convertir el color a minúscula
        number1 = match.group(2)  # Primer número
        number2 = match.group(3) if match.group(3) else ""  # Segundo número opcional
        ubicacion = f"{color_prefix}{number1}{number2}"
    else:
        ubicacion = "no_location"

    return ubicacion

def extract_color_counts(message: str):
    """Extrae las cantidades de colores del mensaje y cuenta las banderas si no se encuentra el patrón específico."""
    
    # Inicialización de los conteos de banderas y el patrón de búsqueda
    color_counts = {'🇲🇴': 0, '🇻🇦': 0, '🇮🇲': 0, '🇪🇺': 0}
    color_patterns = {'🇲🇴': 0, '🇻🇦': 0, '🇮🇲': 0, '🇪🇺': 0}

    # Buscar patrones específicos de colores
    for color in color_counts.keys():
        count_pattern = rf"{color}\s*:\s*(\d+)"
        count_match = re.search(count_pattern, message)
        if count_match:
            color_counts[color] = int(count_match.group(1))

    # Si no se encontraron patrones específicos, contar todas las banderas
    if all(value == 0 for value in color_counts.values()):
        for color in color_counts.keys():
            color_patterns[color] = len(re.findall(color, message))
        color_counts = color_patterns

    return color_counts

def find_row_for_location(ws, location):
    """Busca la fila correspondiente a una ubicación en el Excel."""
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == location:
            return row
    return None

async def get_excel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    #####SEGURIDAD#####
    language_code = update.effective_user.language_code
    user_id = update.effective_user.id
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text(get_text(update, 'no_permission'))
        return []
    ##################
    
    file_path = EXCEL_PATH
        
    try:
        # Envía el documento al usuario
        with open(file_path, 'rb') as excel_file:
            await update.message.reply_document(document=excel_file)
    except FileNotFoundError:
        await update.message.reply_text(get_text(update,'no_excel'))
    except Exception as e:
        await update.message.reply_text(f"A error has ocurred: {e}")

def save_to_excel(ws, location, color_counts, text, user_posted):
    """Guarda los datos extraídos en el archivo Excel."""
    row = find_row_for_location(ws, location)

    if row is None:
        # Si la ubicación no existe, crea una nueva fila
        new_row = [location] + [color_counts.get(color_emoji, 0) for color_emoji in ['🇲🇴', '🇻🇦', '🇮🇲', '🇪🇺']] + [text, time,user_posted]
        ws.append(new_row)
    else:
        # Si la ubicación ya existe, actualiza la fila existente
        for i, color_emoji in enumerate(['🇲🇴', '🇻🇦', '🇮🇲', '🇪🇺'], start=2):
            ws.cell(row=row, column=i).value = color_counts.get(color_emoji, 0)
        ws.cell(row=row, column=6).value = text  # Actualiza el texto en la columna 6
        ws.cell(row=row, column=7).value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws.cell(row=row, column=8).value = user_posted  # Actualiza el nombre del usuario en la columna 8

    # Guardar los cambios en el archivo Excel
    ws.parent.save(EXCEL_PATH)


patterncomp = re.compile(
    r"^You (climbed to the highest point in the|looked to the)\s+"  # Primera línea
    r"(?:0#0|.*?([RGBY]{1,2})[\s\[\]]*(\d+)[\s\[\]]*(?:#(\d+))?.*?)"  # Ubicación (opcional)
    r"(?:Total:\s*\d+\s*👥\s*(?:🇲🇴|🇻🇦|🇮🇲|🇪🇺\s*:\s*\d+\s*👥,\s*Leader:\s*.+\s*)?)?" # Total e información de equipo (opcional)
    r"((?:🇲🇴|🇻🇦|🇮🇲|🇪🇺[\w\d\s]+ 🏅\d+ 👣\d+\s*)*)",  # Lista de usuarios (opcional),
    #r"(?:Combat options: /combat)?$",  # Opción de combate (opcional)
    re.DOTALL | re.MULTILINE
)

def es_mensaje_valido(mensaje: str) -> bool:
    # Si hay información de equipo, debe haber total
    #if ("🇲🇴" in mensaje or "🇻🇦" in mensaje or "🇮🇲" in mensaje or "🇪🇺" in mensaje) and "Total:" not in mensaje:
    #    return False
    return bool(patterncomp.match(mensaje))

async def save_message(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    #####SEGURIDAD#####
    if update.message.chat.type in [update.message.chat.SUPERGROUP, update.message.chat.GROUP]:
        return []
    
    if hasattr(update.message, 'forward_from') and update.message.forward_from and update.message.forward_from.id == 265204902:
        await update.message.reply_text(get_text(update, 'message_forwarded'))
    elif hasattr(update.message, 'forward_origin') and update.message.forward_origin and update.message.forward_origin.sender_user and update.message.forward_origin.sender_user.id == 265204902:
        await update.message.reply_text(get_text(update, 'processing_info'))
    else:
        await update.message.reply_text(get_text(update, 'processing_info') + "NOT")
        return []
    ###################
    
    message = update.message.text
    
    if not es_mensaje_valido(message):
        await update.message.reply_text(get_text(update, 'invalid_message'))
        return []
    
    wb, ws = cargar_o_crear_excel()
    
    ubicacion = extract_location(message)
    color_counts = extract_color_counts(message)
    
    user_posted = update.effective_user.username or update.effective_user.full_name

    save_to_excel(ws, ubicacion, color_counts, message, user_posted)

    msg = ""
    for color_emoji, count in color_counts.items():
        msg += f"\n{color_emoji} -> {count}"
    await update.message.reply_text(
    get_text(update, 'saved_successfully').format(location=ubicacion, msg=msg, user_posted=user_posted)
)



async def send_map(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    #####SEGURIDAD#####
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text(get_text(update, 'no_permission'))
        return []
    ##################
    
    try:
        await update.message.reply_photo(photo=InputFile("map.jpg"))
    except Exception as e:
        await update.message.reply_text(get_text(update, 'error_sending_image').format(error=e))


async def info(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    #####SEGURIDAD#####
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text(get_text(update, 'no_permission'))
        return []
    ##################
    
    wb, ws = cargar_o_crear_excel()
    
    location = context.args[0] if context.args else None
    if not location:
        await update.message.reply_text(get_text(update, 'provide_location'))
        return

    row = find_row_for_location(ws, location)
    
    if row:
        saved_time = ws.cell(row=row, column=7).value
        time_difference = 0
        
        if saved_time:
            saved_time_dt = datetime.strptime(saved_time, '%Y-%m-%d %H:%M:%S')
            local_tz = pytz.timezone('America/Havana')
            saved_time_local = saved_time_dt.astimezone(local_tz)
            current_time = datetime.now(local_tz)
            time_difference = int((current_time - saved_time_local).total_seconds() / 60)
        else:
            time_difference = -1
        
        text = details=ws.cell(row=row, column=6).value
        text += get_text(update, 'simple_info_footer').format(
            time_difference=time_difference,
            user=ws.cell(row=row, column=8).value
        )
        await update.message.reply_text(text)
    else:
        await update.message.reply_text(get_text(update, 'no_info_found').format(location=location))

async def simple_info(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    #####SEGURIDAD#####
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text(get_text(update, 'no_permission'))
        return []
    ##################
    
    wb, ws = cargar_o_crear_excel()
    
    location = context.args[0] if context.args else None
    if not location:
        await update.message.reply_text(get_text(update, 'provide_location'))
        return

    row = find_row_for_location(ws, location)
    color_emoji = ['🇲🇴', '🇻🇦', '🇮🇲', '🇪🇺']
    
    if row:
        saved_time = ws.cell(row=row, column=7).value
        time_difference = 0
        
        if saved_time:
            saved_time_dt = datetime.strptime(saved_time, '%Y-%m-%d %H:%M:%S')
            local_tz = pytz.timezone('America/Havana')
            saved_time_local = saved_time_dt.astimezone(local_tz)
            current_time = datetime.now(local_tz)
            time_difference = int((current_time - saved_time_local).total_seconds() / 60)
        else:
            time_difference = -1
        
        msg = get_text(update, 'simple_info_header').format(location=location)
        color_counts = [ws.cell(row, column=i).value for i in range(2, 6)]
        
        for x, count in enumerate(color_counts):
            msg += get_text(update, 'color_count').format(color_emoji[x], count)
        msg += get_text(update, 'simple_info_footer').format(time_difference=time_difference, user=ws.cell(row=row, column=8).value)
        await update.message.reply_text(msg)
    else:
        await update.message.reply_text(get_text(update, 'no_info_found').format(location=location))



async def help(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    #####SEGURIDAD#####
    language_code = update.effective_user.language_code
    user_id = update.effective_user.id
    if not is_authorized(update.effective_user.id):
        await update.message.reply_text(get_text(update, 'no_permission'))
        return []
    ##################
    # Obtener el mensaje de ayuda basado en el idioma
    await update.message.reply_text(get_text(update, 'help_message'))

# Nuevo comando para establecer el idioma
async def set_language(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Permite al usuario seleccionar su idioma preferido."""
    # Crear botones de idioma
    buttons = [
        [InlineKeyboardButton("English", callback_data='en')],
        [InlineKeyboardButton("Español", callback_data='es')],
        [InlineKeyboardButton("Русский", callback_data='ru')]
    ]
    
    # Crear y enviar el teclado inline
    keyboard = InlineKeyboardMarkup(buttons)
    await update.message.reply_text(get_text(update, 'choose_language'), reply_markup=keyboard)

async def set_language_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    selected_language = query.data
    user_id = query.from_user.id

    # Cambia el idioma del usuario en USER_LANGUAGES
    USER_LANGUAGES[user_id] = selected_language

    # Obtiene el mensaje de confirmación en el nuevo idioma
    confirmation_text = get_text(update, 'choose_language')

    # Verifica si el nuevo texto es diferente del actual antes de editar
    if query.message.text != confirmation_text:
        await query.edit_message_text(text=confirmation_text)
    else:
        await query.answer()  # Esto previene que el callback quede "pendiente"



# Configuración del bot
app = ApplicationBuilder().token("6436295787:AAHQYGQj94g_1iuuzmU5RQa43esNok7Cj3g").build()
# Explorer bot: 6436295787:AAHQYGQj94g_1iuuzmU5RQa43esNok7Cj3g
# Test bot: 7523544789:AAE6u1waeC3kL3LpZK_7-J_CNqNTdPbybG4


app.add_handler(CommandHandler("start", start))
app.add_handler(CommandHandler("help", help))
app.add_handler(CommandHandler("get_excel", get_excel))
app.add_handler(CommandHandler("map", send_map))
app.add_handler(CommandHandler("i", simple_info))  # Comando /i original
app.add_handler(CommandHandler("info", info))   # Nuevo comando /info
app.add_handler(CommandHandler("set_language", set_language))
app.add_handler(CallbackQueryHandler(set_language_callback))
app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, save_message))

# Inicia el bot
app.run_polling()
